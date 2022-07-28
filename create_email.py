import win32com.client
import openpyxl
import os
##from main import flag_arrow_link

dict_email = {};
Separation_Char = None;
EmptySpace_Char = None;
flag_arrow_link = False;

def getExcelData(workbook, sheetName):
    global Separation_Char
    global EmptySpace_Char
    pxl_doc = openpyxl.load_workbook(workbook, data_only = True);
    sheet = pxl_doc[sheetName];
    excelData = {};
    # iterate over columns in the first row
    for i in range(1, sheet.max_column+1):
        excelData[sheet.cell(row=1, column=i).value] = sheet.cell(row=2, column=i).value;

    ## Used for sinchronizaton over variables within the excel file
    sheet = pxl_doc["CONST"];
    Separation_Char = sheet.cell(row = 1, column = 2).value;
    EmptySpace_Char = sheet.cell(row = 2, column = 2).value;

    return excelData

def drawHTMLTable(col_1, col_2, col_3 = None, hyper_1 = None, hyper_2 = None, hyper_3 = None):
    global Separation_Char
    global EmptySpace_Char
    col_1 = col_1.split(Separation_Char);
    col_2 = col_2.split(Separation_Char);

    if type(col_3) == type(None):
        pass
    
    else:
        col_3 = col_3.split(Separation_Char);

    table = [];
    table.append(col_1);
    table.append(col_2);
    table.append(col_3);

    if type(hyper_1) == type(None):
        pass
    
    else:
        hyper_1 = hyper_1.split(Separation_Char);
        hyper_1.insert(0, None);

    if type(hyper_2) == type(None):
        pass
    
    else:
        hyper_2 = hyper_2.split(Separation_Char);
        hyper_2.insert(0, None);

    if type(hyper_3) == type(None):
        pass
    
    else:
        hyper_3 = hyper_3.split(Separation_Char);
        hyper_3.insert(0, None);

    hyper = [];
    hyper.append(hyper_1);
    hyper.append(hyper_2);
    hyper.append(hyper_3);
    
    if type(col_3) == type(None):
        HTML_str = "<table>";
        for i in range(len(table[0])):
            HTML_str = HTML_str + "<tr>";
            if i == 0:
                HTML_str = HTML_str + "<th width = \"70%\">{}&emsp;</th><th>{}</th>".format(table[0][i], table[1][i]);

            else:
                HTML_str = HTML_str + "<td>{}&emsp;</td><td align = \"center\">{}</td>".format(table[0][i], table[1][i]);
                
            HTML_str = HTML_str + "</tr>";

    else:
        HTML_str = "<table width=\"800\">";
        HTML_str = HTML_str + "<style> td {word-wrap: break-word;} </style>";
        for i in range(len(table[0])):
            HTML_str = HTML_str + "<tr>";
            for j in range(len(table)):
                ## Place Headers in first definition of table
                if i == 0:
                    if j == 0:
                        HTML_str = HTML_str + "<th style = \"width: 30%\">";

                    elif j == 1:
                        HTML_str = HTML_str + "<th>";

                    elif j == 2:
                        HTML_str = HTML_str + "<th style = \"width: 25%\">";

                    HTML_str = HTML_str + "{}</th>".format(table[j][i]);
                        
                ## if headers have already been placed, continue as normal
                else:
                    HTML_str = HTML_str + "<td align = \"center\">";
                    
                    ## If no Hyperlink is provided in general, insert as text
                    if type(hyper[j]) == type(None):
                        HTML_str = HTML_str + "{}".format(table[j][i]);

                    else:
                        ## If no Hyperlink is provided for specific cell, insert as text
                        if hyper[j][i] == EmptySpace_Char:
                            HTML_str = HTML_str + "{}".format(table[j][i]);

                        else:
                            HTML_str = HTML_str + "<a href={}>{}".format(hyper[j][i], table[j][i]);
                    
            HTML_str = HTML_str + "</tr>";
            
    HTML_str = HTML_str + "</table>";

    return HTML_str;

def func_genSubmissionEmail(data, pptx_name):
    global flag_arrow_link
    
    data = getExcelData("./NPI_TEMPLATE_FILL_Test.xlsx", 'TEMPLATE_1_FILL')

    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

    msg = outlook.OpenSharedItem("{}\\{}".format(os.getcwd(), "SubmissionEmail_Template.msg"))
##    print(msg.HTMLBody);

    msg.Subject = "{} {} - {}".format(msg.Subject, data['Supplier'], data["Title"]);
    msg.HTMLBody = msg.HTMLBody.replace("<p><span class=SpellE>TableHere</span><o:p></o:p></p>",
                                        "{}".format(drawHTMLTable(col_1 = data['OPNTableColumn1'], col_2 = data['OPNTableColumn4'])));

##    print(msg.HTMLBody);
    msg.HTMLBody = msg.HTMLBody.replace("Supplier:", "Supplier: " + data['Supplier'])
    if flag_arrow_link:
        msg.HTMLBody = msg.HTMLBody.replace("Arrow.com link:", "Arrow.com link:&nbsp;Yes");

    else:
        msg.HTMLBody = msg.HTMLBody.replace("Arrow.com link:", "Arrow.com link:&nbsp;No");
        
    msg.HTMLBody = msg.HTMLBody.replace("Topic:", "Topic: " + data['Title'])
    msg.HTMLBody = msg.HTMLBody.replace("Application:", "Application:\n" + data['ApplicationText'])

    msg.Attachments.Add("{}\\{}".format(os.getcwd(), pptx_name));
    msg.Attachments.Add("{}\\{}".format(os.getcwd(), "{}.msg".format(pptx_name[0:pptx_name.find(".")])));

    msg_name = "{}\\{}".format(os.getcwd(), "Submission-{}.msg".format(pptx_name[0:pptx_name.find(".")]));
    msg.SaveAs(msg_name, 3);
    print("PROGRESS: Saving <{}>".format(msg_name));

    del msg, outlook
    
    try:
        outlook = win32com.client.Dispatch('outlook.application')
        mail.outlook.CreateItem(0);
        mail.To = "ESCAmericas@arrow.com"
        mail.Subject = "DONAS"
        mail.Body = "Donas, todos, amigos"
        mail.Send();

        del mail, outlook

    except:
        pass

    

def func_FMTText(text, fmt_chr):
    flag = True;
    
    while True:
        if text.find(fmt_chr) >= 0:
            if flag:
                text = text.replace(fmt_chr, "<b>", 1);
                flag = False;

            else:
                text = text.replace(fmt_chr, "</b>", 1)
                flag = True;

        else:
            break;
        
    return text;
    

def func_genClientEmail(data, pptx_name):
    data = getExcelData("./NPI_TEMPLATE_FILL_Test.xlsx", 'TEMPLATE_1_FILL')

    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

    msg = outlook.OpenSharedItem("{}\\{}".format(os.getcwd(), "CustomerEmail_Template.msg"))

    msg.Subject = "{} - {}".format(data['Supplier'], data["Title"]);

    msg.Attachments.Add("{}\\{}".format(os.getcwd(), pptx_name));
    
    if data['EmailText'].find("__") >= 0:
        if data['EmailText'].count("__") % 2:
##            main.printWARN("WARNING: Bold chars <\"{}\"> are not multiple of 2 in <{}>, proceeding without format".format(dict_fmt['bold'],aux));
            
            msg.HTMLBody = msg.HTMLBody.replace("EmailText", data['EmailText'].replace("__", ""))

        else:
            msg.HTMLBody = msg.HTMLBody.replace("EmailText", func_FMTText(data['EmailText'], "__"))
   
    msg.HTMLBody = msg.HTMLBody.replace("AdvantagesText", data['AdvantagesText'])
    msg.HTMLBody = msg.HTMLBody.replace("ApplicationText", data['ApplicationText'])

    width_index = msg.HTMLBody.rfind("width=") + len("width=");
    width_str = "";
    while True:
        if msg.HTMLBody[width_index] == " ":
            break;

        width_str = width_str + msg.HTMLBody[width_index];
        width_index += 1;
    
    height_index = msg.HTMLBody.rfind("height=") + len("height=");
    height_str = "";
    while True:
        if msg.HTMLBody[height_index] == " ":
            break;

        height_str = height_str + msg.HTMLBody[height_index];
        height_index += 1;

    main_dir = os.getcwd();
    
    print(main_dir);
    attachment1 = msg.Attachments.Add("{}\\images\\figures\\{}".format(main_dir, data['MainFigureImage']))
    attachment1.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", "imgId1")

    attachment2 = msg.Attachments.Add("{}\\images\\figures\\{}".format(main_dir, data['SubFigure1Image']))
    attachment2.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", "imgId2")
    
    msg.HTMLBody = msg.HTMLBody.replace("<img width={} height={} id=\"Picture_x0020_2\" src=\"cid:image001.jpg@01D89A90.88E97900\" alt=\"Image Placeholder - Uning\">".format(width_str, height_str),
                                        "<img height={} src=\'cid:{}\' alt=\'{}\'>".format(height_str, "imgId1", data['Supplier']))
    msg.HTMLBody = msg.HTMLBody.replace("<img width={} height={} id=\"Picture_x0020_1\" src=\"cid:image001.jpg@01D89A90.88E97900\" alt=\"Image Placeholder - Uning\">".format(width_str, height_str),
                                        "<img height={} src=\'cid:{}\' alt=\'{}\'>".format(height_str, "imgId2", data['Supplier']))
    
    msg.HTMLBody = msg.HTMLBody.replace("<span style='color:black'>TableHere</span>",
                                        "{}".format(drawHTMLTable(col_1 = data['OPNTableColumn1'], col_2 = data['OPNTableColumn2'], col_3 = data['OPNTableColumn3'], hyper_1 = data['OPNTableColumn1Link'], hyper_3 = data['OPNTableColumn3Link'])));

    if not type(data['PageLink1']) == type(None) and not type(data['PageLink2']) == type(None):
        msg.HTMLBody = msg.HTMLBody.replace("<span style='color:black'>PageLink1|PageLink2<o:p></o:p></span>",
                                            "<a href={}>{}</a><FONT FACE=\"Calibri\"> |  </FONT><a href={}>{}</a>".format(data['PageLink1'], data['PageLink1Mask'], data['PageLink2'], data['PageLink2Mask']))

    elif not type(data['PageLink1']) == type(None):
        msg.HTMLBody = msg.HTMLBody.replace("<span style='color:black'>PageLink1|PageLink2<o:p></o:p></span>",
                                            "<font><&nbsp;</font><a href={}>{}</a><font>&nbsp;></font>".format(data['PageLink1'], data['PageLink1Mask']))

    elif not type(data['PageLink2']) == type(None):
        msg.HTMLBody = msg.HTMLBody.replace("<span style='color:black'>PageLink1|PageLink2<o:p></o:p></span>",
                                            "<font><&nbsp;</font><a href={}>{}</a><font>&nbsp;></font>".format(data['PageLink2'], data['PageLink2Mask']))

    else:
        msg.HTMLBody = msg.HTMLBody.replace("<span style='color:black'>PageLink1|PageLink2<o:p></o:p></span>","");


##    print(msg.HTMLBody);

    msg_name = "{}\\{}".format(os.getcwd(), "{}.msg".format(pptx_name[0:pptx_name.find(".")]));
    msg.SaveAs(msg_name, 3);
    print("PROGRESS: Saving <{}>".format(msg_name));

    del msg, outlook
